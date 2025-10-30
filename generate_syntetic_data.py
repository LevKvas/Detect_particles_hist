import numpy as np
import pandas as pd
import os
from openpyxl import Workbook


def read_real_data(file_path):
    """Читает реальные данные из txt-файла и возвращает гистограммы"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
    except FileNotFoundError:
        print(f"Файл {file_path} не найден!")
        return [], []

    eccentricity_data = []
    square_data = []
    section = None

    for line in lines:
        line = line.strip()
        if "ECCENTRICITY HISTOGRAM" in line:
            section = 'eccentricity'
            continue
        elif "SQUARE HISTOGRAM" in line:
            section = 'square'
            continue
        elif "Bin_Center" in line or "-----" in line or not line or "HISTOGRAM DATA" in line or "====" in line:
            continue

        if section == 'eccentricity' and line:
            parts = line.split()
            if len(parts) == 2:
                try:
                    eccentricity_data.append((float(parts[0]), int(parts[1])))
                except ValueError:
                    continue
        elif section == 'square' and line:
            parts = line.split()
            if len(parts) == 2:
                try:
                    square_data.append((float(parts[0]), int(parts[1])))
                except ValueError:
                    continue

    return eccentricity_data, square_data


def generate_synthetic_histogram(real_data, num_particles=1000):
    """Генерирует синтетическую гистограмму на основе реального распределения"""
    if not real_data:
        return []

    bin_centers, counts = zip(*real_data)

    # Создаем распределение вероятностей на основе счетчиков
    total_count = sum(counts)
    if total_count == 0:
        return []

    probabilities = np.array(counts) / total_count

    # Генерируем синтетические частицы
    synthetic_counts = np.zeros(len(bin_centers), dtype=int)

    for _ in range(num_particles):
        # Выбираем случайный бин согласно распределению вероятностей
        bin_idx = np.random.choice(len(bin_centers), p=probabilities)
        synthetic_counts[bin_idx] += 1

    # Возвращаем в том же формате, что и исходные данные
    return list(zip(bin_centers, synthetic_counts))


def create_synthetic_distributions(num_distributions=5, particles_per_distribution=1000):
    """Создает синтетические распределения на основе реальных данных"""

    dir_file = "exist_hists/"
    real_files = ['3_torr_new.txt', '48_torr_new.txt', '72_torr_new.txt',
                  '96_torr_new.txt', '196_torr_new.txt']
    for i in range(len(real_files)):
        real_files[i] = dir_file + real_files[i]

    # Проверяем, какие файлы существуют
    available_files = [f for f in real_files if os.path.exists(f)]

    if not available_files:
        print("Ни один из файлов не найден! Убедитесь, что файлы находятся в той же папке.")
        return

    print(f"Найдено файлов: {len(available_files)}")

    # Создаем Excel файл
    output_file = 'synthetic_distributions.xlsx'

    if os.path.exists(output_file):
        os.remove(output_file)

    # Создаем новую книгу Excel
    wb = Workbook()
    # Удаляем дефолтный лист, если он есть
    if wb.sheetnames:
        wb.remove(wb.active)

    distributions_created = 0

    for dist_num in range(1, num_distributions + 1):
        # Случайно выбираем базовый файл для этого распределения
        base_file = np.random.choice(available_files)

        # Читаем реальные данные
        ecc_data, sq_data = read_real_data(base_file)

        # Проверяем, что данные не пустые
        if not ecc_data or not sq_data:
            print(f"Пропускаем распределение {dist_num}: нет данных в файле {base_file}")
            continue

        # Генерируем синтетические гистограммы
        synthetic_ecc = generate_synthetic_histogram(ecc_data, particles_per_distribution)
        synthetic_sq = generate_synthetic_histogram(sq_data, particles_per_distribution)

        # Проверяем, что синтетические данные созданы
        if not synthetic_ecc or not synthetic_sq:
            print(f"Пропускаем распределение {dist_num}: не удалось сгенерировать данные")
            continue

        # Создаем DataFrame для эксцентриситета
        ecc_bins, ecc_counts = zip(*synthetic_ecc)
        df_ecc = pd.DataFrame({
            'Bin_Center': ecc_bins,
            'Particle_Count': ecc_counts
        })

        # Создаем DataFrame для площади
        sq_bins, sq_counts = zip(*synthetic_sq)
        df_sq = pd.DataFrame({
            'Bin_Center': sq_bins,
            'Particle_Count': sq_counts
        })

        # Создаем новый лист
        sheet_name = f"Dist_{dist_num}"
        ws = wb.create_sheet(title=sheet_name)

        # Записываем заголовки и данные
        ws.cell(1, 1, "ECCENTRICITY HISTOGRAM")
        # Записываем данные эксцентриситета
        ws.cell(3, 1, "Bin_Center")
        ws.cell(3, 2, "Particle_Count")

        for idx, (bin_center, count) in enumerate(synthetic_ecc, start=4):
            ws.cell(idx, 1, bin_center)
            ws.cell(idx, 2, count)

        # Записываем заголовок площади
        start_row = len(synthetic_ecc) + 6
        ws.cell(start_row, 1, "SQUARE HISTOGRAM (nm²)")
        ws.cell(start_row + 1, 1, "Bin_Center")
        ws.cell(start_row + 1, 2, "Particle_Count")

        # Записываем данные площади
        for idx, (bin_center, count) in enumerate(synthetic_sq, start=start_row + 2):
            ws.cell(idx, 1, bin_center)
            ws.cell(idx, 2, count)

        # Добавляем информацию о распределении
        info_row = start_row + len(synthetic_sq) + 4
        ws.cell(info_row, 1, f"Based on: {base_file}")
        ws.cell(info_row + 1, 1, f"Total particles: {particles_per_distribution}")

        distributions_created += 1
        print(f"Создано распределение {dist_num} на основе {base_file}")

    # Сохраняем файл только если созданы распределения
    if distributions_created > 0:
        wb.save(output_file)
        print(f"Успешно создано {distributions_created} синтетических распределений в файле '{output_file}'")
    else:
        print("Не удалось создать ни одного распределения!")


# Дополнительная функция: создать распределения на основе конкретного файла
def create_synthetic_from_specific_file(base_file, num_distributions=5, particles_per_distribution=1000):
    """Создает синтетические распределения на основе конкретного файла"""

    if not os.path.exists(base_file):
        print(f"Файл {base_file} не найден!")
        return

    output_file = f'synthetic_from_{base_file.replace(".txt", "")}.xlsx'

    if os.path.exists(output_file):
        os.remove(output_file)

    # Создаем новую книгу Excel
    wb = Workbook()
    # Удаляем дефолтный лист
    if wb.sheetnames:
        wb.remove(wb.active)

    distributions_created = 0

    for dist_num in range(1, num_distributions + 1):
        # Читаем реальные данные
        ecc_data, sq_data = read_real_data(base_file)

        # Проверяем, что данные не пустые
        if not ecc_data or not sq_data:
            print(f"Пропускаем распределение {dist_num}: нет данных в файле {base_file}")
            continue

        # Генерируем синтетические гистограммы
        synthetic_ecc = generate_synthetic_histogram(ecc_data, particles_per_distribution)
        synthetic_sq = generate_synthetic_histogram(sq_data, particles_per_distribution)

        # Проверяем, что синтетические данные созданы
        if not synthetic_ecc or not synthetic_sq:
            print(f"Пропускаем распределение {dist_num}: не удалось сгенерировать данные")
            continue

        # Создаем DataFrame
        ecc_bins, ecc_counts = zip(*synthetic_ecc)
        sq_bins, sq_counts = zip(*synthetic_sq)

        # Создаем новый лист
        sheet_name = f"Dist_{dist_num}"
        ws = wb.create_sheet(title=sheet_name)

        # Записываем данные эксцентриситета
        ws.cell(1, 1, "ECCENTRICITY HISTOGRAM")
        ws.cell(3, 1, "Bin_Center")
        ws.cell(3, 2, "Particle_Count")

        for idx, (bin_center, count) in enumerate(synthetic_ecc, start=4):
            ws.cell(idx, 1, bin_center)
            ws.cell(idx, 2, count)

        # Записываем данные площади
        start_row = len(synthetic_ecc) + 6
        ws.cell(start_row, 1, "SQUARE HISTOGRAM (nm²)")
        ws.cell(start_row + 1, 1, "Bin_Center")
        ws.cell(start_row + 1, 2, "Particle_Count")

        for idx, (bin_center, count) in enumerate(synthetic_sq, start=start_row + 2):
            ws.cell(idx, 1, bin_center)
            ws.cell(idx, 2, count)

        # Добавляем информацию
        info_row = start_row + len(synthetic_sq) + 4
        ws.cell(info_row, 1, f"Based on: {base_file}")
        ws.cell(info_row + 1, 1, f"Total particles: {particles_per_distribution}")

        distributions_created += 1

    # Сохраняем файл
    if distributions_created > 0:
        wb.save(output_file)
        print(f"Создано {distributions_created} синтетических распределений на основе {base_file}")
    else:
        print(f"Не удалось создать распределения на основе {base_file}")


def analyze_real_data():
    """Анализирует реальные данные для понимания масштабов"""
    real_files = ['3_torr_new.txt', '48_torr_new.txt', '72_torr_new.txt',
                  '96_torr_new.txt', '196_torr_new.txt']

    total_particles_list = []

    for file in real_files:
        if os.path.exists(file):
            ecc_data, sq_data = read_real_data(file)

            total_ecc_particles = sum(count for _, count in ecc_data)
            total_sq_particles = sum(count for _, count in sq_data)

            print(f"{file}:")
            print(f"  Всего частиц по эксцентриситету: {total_ecc_particles}")
            print(f"  Всего частиц по площади: {total_sq_particles}")

            total_particles_list.append(total_ecc_particles)

    if total_particles_list:
        avg_particles = np.mean(total_particles_list)
        std_particles = np.std(total_particles_list)
        print(f"\nСреднее количество частиц: {avg_particles:.1f} ± {std_particles:.1f}")

    return total_particles_list


def get_realistic_particle_count(base_file):
    """Возвращает реалистичное количество частиц на основе реальных данных"""
    if os.path.exists(base_file):
        ecc_data, _ = read_real_data(base_file)
        real_count = sum(count for _, count in ecc_data)

        # Добавляем небольшой случайный разброс (±20%)
        variation = np.random.normal(0, 0.1)  # 10% вариация
        synthetic_count = int(real_count * (1 + variation))

        # Ограничиваем разумными пределами
        synthetic_count = max(20, min(200, synthetic_count))

        return synthetic_count
    else:
        # Если файл не найден, используем среднее из реальных данных
        return 50


def create_realistic_synthetic_distributions(num_distributions=5):
    """Создает реалистичные синтетические распределения"""

    dir_file = "exist_hists/"
    real_files = ['3_torr_new.txt', '48_torr_new.txt', '72_torr_new.txt',
                  '96_torr_new.txt', '196_torr_new.txt']
    for i in range(len(real_files)):
        real_files[i] = dir_file + real_files[i]

    available_files = [f for f in real_files if os.path.exists(f)]

    if not available_files:
        print("Ни один из файлов не найден!")
        return

    print("Анализ реальных данных:")
    analyze_real_data()

    output_file = 'realistic_synthetic_distributions.xlsx'

    if os.path.exists(output_file):
        os.remove(output_file)

    wb = Workbook()
    if wb.sheetnames:
        wb.remove(wb.active)

    distributions_created = 0

    for dist_num in range(1, num_distributions + 1):
        base_file = np.random.choice(available_files)
        ecc_data, sq_data = read_real_data(base_file)

        if not ecc_data or not sq_data:
            continue

        # Используем реалистичное количество частиц!
        realistic_particles = get_realistic_particle_count(base_file)

        synthetic_ecc = generate_synthetic_histogram(ecc_data, realistic_particles)
        synthetic_sq = generate_synthetic_histogram(sq_data, realistic_particles)

        if not synthetic_ecc or not synthetic_sq:
            continue

        # Создаем лист
        sheet_name = f"Dist_{dist_num}"
        ws = wb.create_sheet(title=sheet_name)

        # Записываем данные (как в предыдущем коде)
        ws.cell(1, 1, "ECCENTRICITY HISTOGRAM")
        ws.cell(3, 1, "Bin_Center")
        ws.cell(3, 2, "Particle_Count")

        for idx, (bin_center, count) in enumerate(synthetic_ecc, start=4):
            ws.cell(idx, 1, bin_center)
            ws.cell(idx, 2, count)

        start_row = len(synthetic_ecc) + 6
        ws.cell(start_row, 1, "SQUARE HISTOGRAM (nm²)")
        ws.cell(start_row + 1, 1, "Bin_Center")
        ws.cell(start_row + 1, 2, "Particle_Count")

        for idx, (bin_center, count) in enumerate(synthetic_sq, start=start_row + 2):
            ws.cell(idx, 1, bin_center)
            ws.cell(idx, 2, count)

        # Добавляем реалистичную информацию
        info_row = start_row + len(synthetic_sq) + 4
        real_count = sum(count for _, count in ecc_data)
        ws.cell(info_row, 1, f"Based on: {base_file}")
        ws.cell(info_row + 1, 1, f"Real particles in source: {real_count}")
        ws.cell(info_row + 2, 1, f"Synthetic particles: {realistic_particles}")

        distributions_created += 1
        print(f"Создано распределение {dist_num}: {realistic_particles} частиц (источник: {base_file})")

    if distributions_created > 0:
        wb.save(output_file)
        print(f"\nСоздано {distributions_created} реалистичных распределений")


# Пример использования:
if __name__ == "__main__":
    print("Создание РЕАЛИСТИЧНЫХ синтетических данных...")

    # Создать 100 реалистичных распределения
    create_realistic_synthetic_distributions(num_distributions=100)

